from openpyxl import *
from openpyxl.styles import Font

from injury_record import InjuryRecord

def save_record(record: InjuryRecord):

    wb = Workbook()
    ws = wb.active
    ws.title = "Injury Data"

    columns = ["ID", "Type", "Locations", "Area", "Note"]

    # add column names
    ws.append(columns)

    bold_font = Font(bold=True)
    for cell in ws[1]:  # ws[1] refers to the first row
        cell.font = bold_font

    for injury in record.injury_list:
        ws.append([injury.id, injury.type, injury.get_locations_string(), injury.area, injury.note])

    wb.save(f"client-data/{record.client}-{record.safe_date_format()}.xlsx")
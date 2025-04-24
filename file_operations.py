from shutil import copyfile

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment, Protection
from pathlib import Path
from pydrive2.auth import GoogleAuth
from pydrive2.drive import GoogleDrive
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime

from data.body_map_data import find_secondary_range_side
from injury_record import InjuryRecord
import os

# path to service account JSON key file
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SERVICE_ACCOUNT_FILE = os.path.join(BASE_DIR, "drive_creds", "sharp-sandbox-455617-e9-76fea66cb16c.json")

# the Google Drive folder ID
FOLDER_ID = "1EqrtKzLiKc2UpjE_EFJJo5bVGEEhXbMY"


def authenticate_drive():
    gauth = GoogleAuth()
    scope = ["https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name(SERVICE_ACCOUNT_FILE, scope)
    gauth.credentials = creds
    return GoogleDrive(gauth)


# authenticate + create Google Drive instance
drive = authenticate_drive()

# list files inside the specific folder
query = f"'{FOLDER_ID}' in parents and trashed=false"
file_list = drive.ListFile({'q': query}).GetList()


# debug print file names and IDs
# for file in file_list:
#     print(f"Title: {file['title']}, ID: {file['id']}")


def initialize_master_sheet(filepath):
    pass


def get_or_create_drive_folder(service, parent_id, folder_name):
    query = f"'{parent_id}' in parents and trashed = false and title = '{folder_name}' and mimeType = 'application/vnd.google-apps.folder'"
    folder_list = service.ListFile({'q': query}).GetList()

    if folder_list:
        return folder_list[0]['id']
    else:
        new_folder = service.CreateFile({
            'title': folder_name,
            'mimeType': 'application/vnd.google-apps.folder',
            'parents': [{'id': parent_id}]
        })
        new_folder.Upload()
        return new_folder['id']


def save_record(record: InjuryRecord):
    file_name = f"{record.client} {record.safe_date_format()} {record.safe_time_format()}.xlsx"

    # copyfile("excel_templates/template.xlsx", file_name)
    wb = load_workbook("excel_templates/template.xlsx")
    ws = wb["Injury Data"]

    template_row = ws[4]
    ws.delete_rows(4)

    # for injury in record.injury_list:  # need to handle secondary injury area
    #     ws.append([injury.id, injury.type, injury.get_locations_string(), injury.area, injury.note])

    for i, injury in enumerate(record.injury_list):
        row_idx = 4 + i
        ws.insert_rows(row_idx)

        # Apply formatting from the template row to this row
        for col_idx, template_cell in enumerate(template_row, start=1):
            new_cell = ws.cell(row=row_idx, column=col_idx)

            # Copy the font attributes manually
            if template_cell.font:
                new_cell.font = Font(
                    name=template_cell.font.name,
                    size=template_cell.font.size,
                    bold=template_cell.font.bold,
                    italic=template_cell.font.italic,
                    color=template_cell.font.color
                )

            # Copy the border attributes manually
            if template_cell.border:
                new_cell.border = Border(
                    left=template_cell.border.left,
                    right=template_cell.border.right,
                    top=template_cell.border.top,
                    bottom=template_cell.border.bottom,
                    diagonal=template_cell.border.diagonal,
                    diagonal_direction=template_cell.border.diagonal_direction,
                    outline=template_cell.border.outline,
                    vertical=template_cell.border.vertical,
                    horizontal=template_cell.border.horizontal
                )

            # Copy the fill attributes manually
            if template_cell.fill:
                new_cell.fill = PatternFill(
                    start_color=template_cell.fill.start_color,
                    end_color=template_cell.fill.end_color,
                    fill_type=template_cell.fill.fill_type
                )

            # Copy the number_format attributes manually
            if template_cell.number_format:
                new_cell.number_format = template_cell.number_format

            # Copy the protection attributes manually
            if template_cell.protection:
                new_cell.protection = Protection(locked=template_cell.protection.locked,
                                                 hidden=template_cell.protection.hidden)

            # Copy the alignment attributes manually
            if template_cell.alignment:
                new_cell.alignment = Alignment(
                    horizontal=template_cell.alignment.horizontal,
                    vertical=template_cell.alignment.vertical,
                    text_rotation=template_cell.alignment.text_rotation,
                    wrap_text=template_cell.alignment.wrap_text,
                    shrink_to_fit=template_cell.alignment.shrink_to_fit,
                    indent=template_cell.alignment.indent
                )

        # Write injury data
        # need to get secondary injury data

        secondary_info_list = find_secondary_range_side(injury.indices.pop())

        ws.cell(row=row_idx, column=2, value=injury.get_locations_string())
        ws.cell(row=row_idx, column=3, value=secondary_info_list[0])
        ws.cell(row=row_idx, column=4, value=secondary_info_list[1])
        ws.cell(row=row_idx, column=5, value=injury.type)
        ws.cell(row=row_idx, column=6, value=injury.area)
        ws.cell(row=row_idx, column=7, value=injury.note)

    # client info at top of record

    ws["B2"] = f"Client: {record.client}"
    ws["C2"] = f"Date: {record.date}"
    ws["D2"] = f"Time: {record.time}"

    # Summary formulas — added after the last data row
    last_row = 4 + len(record.injury_list) - 1
    ws["J3"] = f"=SUM(F4:F{last_row})"
    ws["J2"] = f"=AVERAGE(F4:F{last_row})"

    # SAVING TO DRIVE

    record_date = datetime.strptime(record.date, "%m/%d/%Y")
    year_str = record_date.strftime("%Y")
    month_str = record_date.strftime("%B")  # e.g., 'April'
    week_str = f"Week {record_date.isocalendar()[1]}"  # ISO week number

    client_initials_dict = get_client_initials_dict()
    client_folder_id = client_initials_dict[record.client]

    # Get or create Year > Month > Week subfolders
    year_folder_id = get_or_create_drive_folder(drive, client_folder_id, year_str)
    month_folder_id = get_or_create_drive_folder(drive, year_folder_id, month_str)
    week_folder_id = get_or_create_drive_folder(drive, month_folder_id, week_str)

    # Save locally
    temp_path = file_name
    wb.save(temp_path)

    # Upload to correct Google Drive folder
    file_drive = drive.CreateFile({'title': file_name, 'parents': [{'id': week_folder_id}]})
    file_drive.SetContentFile(temp_path)
    file_drive.Upload()

    # Clean up
    Path(temp_path).unlink()

    # locate the client-master.xlsx file in Google Drive
    master_file_list = drive.ListFile({
        'q': f"'{client_folder_id}' in parents and title = '{record.client}-master.xlsx' and trashed=false"
    }).GetList()

    if not master_file_list:
        print(f"Error: {record.client}-master.xlsx not found in Google Drive.")
        return

    master_file = master_file_list[0]  # assuming only one master file exists

    # download the master file
    master_temp_path = f"{record.client}-master.xlsx"
    master_file.GetContentFile(master_temp_path)  # download it locally

    # edit master sheet
    master_wb = load_workbook(master_temp_path)
    master_ws = master_wb["Master Data"]


    # Step 1: Find the next empty row in column B after row 3
    row_idx = 4
    while master_ws.cell(row=row_idx, column=2).value is not None:
        row_idx += 1
    row_idx += 1
    # Step 2: Copy styles from the previous row (row_idx - 1)
    for col in range(2, 12):  # Columns B (2) to K (10)
        source_cell = master_ws.cell(row=row_idx - 1, column=col)
        target_cell = master_ws.cell(row=row_idx, column=col)

        target_cell.font = source_cell.font.copy()
        target_cell.fill = source_cell.fill.copy()
        target_cell.border = source_cell.border.copy()
        target_cell.alignment = source_cell.alignment.copy()
        target_cell.number_format = source_cell.number_format
        target_cell.protection = source_cell.protection.copy()

    # Step 3: Insert data into the newly styled row
    date_time = f"{record.get_date()}   {record.time}"
    num_injuries = record.get_num_injuries()
    total_area = record.get_total_injury_area()
    avg_injury_size = record.get_avg_injury_area()
    largest_injury_size = record.get_largest_injury_size()
    injury_type_dict = record.get_injury_type_dict()

    num_bruises = injury_type_dict["Bruise"]
    num_open_wounds = injury_type_dict["Open Wound"]
    num_closed_wounds = injury_type_dict["Closed Wound"]
    num_redness = injury_type_dict["Redness"]
    num_other = injury_type_dict["Other"]

    row_idx -= 1

    master_ws.cell(row=row_idx, column=2, value=date_time)
    master_ws.cell(row=row_idx, column=3, value=num_injuries)
    master_ws.cell(row=row_idx, column=4, value=total_area)
    master_ws.cell(row=row_idx, column=5, value=avg_injury_size)
    master_ws.cell(row=row_idx, column=6, value=largest_injury_size)
    master_ws.cell(row=row_idx, column=7, value=num_bruises)
    master_ws.cell(row=row_idx, column=8, value=num_open_wounds)
    master_ws.cell(row=row_idx, column=9, value=num_closed_wounds)
    master_ws.cell(row=row_idx, column=10, value=num_redness)
    master_ws.cell(row=row_idx, column=11, value=num_other)

    master_wb.save(master_temp_path)  # Save updates

    # re-upload the updated master file
    master_file.SetContentFile(master_temp_path)
    master_file.Upload()

    # remove the temporary master file
    Path(master_temp_path).unlink()


def get_client_data_folder_id():
    # List files at the root level and find the client-data folder
    file_list = drive.ListFile({'q': "title = 'client-data' and trashed = false"}).GetList()

    for file in file_list:
        if file['mimeType'] == 'application/vnd.google-apps.folder':
            return file['id']
    return None


def get_client_initials_dict():
    # Get the client-data folder ID dynamically
    client_data_folder_id = get_client_data_folder_id()

    if not client_data_folder_id:
        print("Error: 'client-data' folder not found.")
        return {}

    # Initialize the dictionary to hold client initials and their corresponding folder ID
    client_initials_dict = {}

    # List all files and folders inside the client-data folder
    file_list = drive.ListFile({'q': f"'{client_data_folder_id}' in parents and trashed=false"}).GetList()

    # Iterate through the files and add client initials folder to the dictionary
    for file in file_list:
        if file['mimeType'] == 'application/vnd.google-apps.folder':
            client_initials_dict[file['title']] = file['id']

    # Print the dictionary for debugging
    print(client_initials_dict)

    return client_initials_dict


def check_for_client(client_initials):
    # Get the client-data folder ID dynamically
    client_data_folder_id = get_client_data_folder_id()

    if not client_data_folder_id:
        print("Error: 'client-data' folder not found.")
        return

    # Check if the client subfolder (e.g., ABC) exists within the client-data folder
    client_folder = None
    file_list = drive.ListFile({'q': f"'{client_data_folder_id}' in parents and trashed=false"}).GetList()
    for file in file_list:
        if file['title'] == client_initials and file['mimeType'] == 'application/vnd.google-apps.folder':
            client_folder = file
            break

    # If the client folder doesn't exist, create it
    if not client_folder:
        client_folder = drive.CreateFile({'title': client_initials, 'mimeType': 'application/vnd.google-apps.folder',
                                          'parents': [{'id': client_data_folder_id}]})
        client_folder.Upload()

    # Create the "client-master.xlsx" file in the client's folder
    template_local_path = "excel_templates/master-template.xlsx"
    master_file_path = f"{client_initials}-master.xlsx"

    # Check if the master file already exists in the client folder
    master_file = None
    file_list = drive.ListFile({'q': f"'{client_folder['id']}' in parents and trashed=false"}).GetList()
    for file in file_list:
        if file['title'] == master_file_path:
            master_file = file
            break

    # If not found, upload a copy of the local template file to the client's folder
    if not master_file:
        new_file = drive.CreateFile({'title': master_file_path, 'parents': [{'id': client_folder['id']}]})
        new_file.SetContentFile(template_local_path)
        new_file.Upload()

